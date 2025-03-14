from datetime import date, datetime
from decimal import Decimal, InvalidOperation
import json
from pyexpat.errors import messages
from django.contrib import messages
from django.db import IntegrityError
from django.shortcuts import get_object_or_404, render, redirect
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from locais.models import Obra, Escritorio
from locais.views import calcular_range_meses, formatar_valor
from .models import BM, Adiantamento, Aditivo, Despesa, Cartao, Fatura, Funcionario, NotaBoleto, NotaPix, NotaEspecie, NotaCartao, MaoDeObra, Banco, Pagamento, Parcela
from django.contrib.contenttypes.models import ContentType
from django.db import transaction
from dateutil.relativedelta import relativedelta
from django.db.models import Prefetch
import cloudinary.uploader
from cloudinary.utils import cloudinary_url


import logging

logger = logging.getLogger(__name__)

def limpar_e_converter_valor(valor):
    try:
        valor = valor.replace('R$', '').replace('.', '').replace(',', '.').strip()
        return Decimal(valor)
    except (ValueError, InvalidOperation):
        raise ValueError("Valor inválido.") 

def limpar_e_converter_data(data_str, request, redirect_url='locais:home', formato='%d/%m/%Y'):
    if not isinstance(data_str, str):
        messages.error(request, 'Data fornecida não é válida ou está vazia.')
        return redirect(redirect_url)
    
    try:
        return datetime.strptime(data_str, formato).date()
    except ValueError:
        messages.error(request, f'Formato de data inválido. Use o formato {formato}.')
        return redirect(redirect_url)

def pegar_escritorio_id(despesa, tipo):
    id_local = despesa.id_local  # Obtém o ID real do local

    if tipo == 'obra':
        local = get_object_or_404(Obra, id=id_local)
        return local.escritorio.id
    elif tipo == 'escritorio':
        local = get_object_or_404(Escritorio, id=id_local)
        return local.id
    else:
        logger.error(f"Tipo inválido fornecido: {tipo}")

# Despesas
@login_required
def criar_despesa(request, tipo, id):
    next_url = request.GET.get('next')
    if tipo == 'obra':
        local = get_object_or_404(Obra, id=id)
        despesas = Despesa.objects.filter(tipo_local=ContentType.objects.get_for_model(local), id_local=id)
        escritorio_id = local.escritorio.id
    elif tipo == 'escritorio':
        local = get_object_or_404(Escritorio, id=id)
        despesas = Despesa.objects.filter(tipo_local=ContentType.objects.get_for_model(local), id_local=id)
        escritorio_id = local.id
    else:
        logger.error(f"Tipo inválido fornecido: {tipo}")
        return redirect('locais:home', escritorio_id=escritorio_id)

    if request.method == 'POST':
        try:
            with transaction.atomic():
                descricao = request.POST.get('descricao')
                forma_pag = request.POST.get('forma_pag')
                status = request.POST.get('status')
                valor = request.POST.get('valor')
                observacao = request.POST.get('observacao')

                data = request.POST.get('data_emissao')
                data_pagamento = request.POST.get('data_pagamento')

                try:
                    if data:
                        data = datetime.strptime(data, '%d/%m/%Y').date()
                    else:
                        messages.error(request, "A data de emissão é obrigatória.")
                        return redirect('locais:home', escritorio_id=escritorio_id)

                    if data_pagamento:
                        data_pagamento = datetime.strptime(data_pagamento, '%d/%m/%Y').date()
                    else:
                        data_pagamento = None
                except ValueError as e:
                    messages.error(request, f'Erro na data: {str(e)}')
                    return redirect('locais:home', escritorio_id=escritorio_id)

                tipo_local = ContentType.objects.get_for_model(local)
                valor = limpar_e_converter_valor(valor)

                # Definindo a modalidade corretamente
                modalidade = request.POST.get('modalidade') if tipo == 'obra' else None
                modalidade_escritorio = request.POST.get('modalidade_escritorio') if tipo == 'escritorio' else None

                # **Validação de modalidade**
                if tipo == 'obra' and not modalidade:
                    messages.error(request, "A modalidade é obrigatória para despesas de Obra.")
                    return redirect('locais:home', escritorio_id=escritorio_id)
                if tipo == 'escritorio' and not modalidade_escritorio:
                    messages.error(request, "A modalidade de escritório é obrigatória para despesas de Escritório.")
                    return redirect('locais:home', escritorio_id=escritorio_id)

                campos_despesa = {
                    "nome": descricao,
                    "forma_pag": forma_pag,
                    "data": data,
                    "valor": valor,
                    "observacao": observacao,
                    "status": status,
                    "tipo_local": tipo_local,
                    "id_local": id,
                    "local": local,
                    "data_pagamento": data_pagamento,
                }

                # Adicionando apenas a modalidade correta
                if tipo == 'obra':
                    campos_despesa["modalidade"] = modalidade
                elif tipo == 'escritorio':
                    campos_despesa["modalidadeEscritorio"] = modalidade_escritorio

                # **Criação da despesa conforme forma de pagamento**
                if forma_pag == 'cartao':
                    cartao_id = request.POST.get('cartao')
                    quant_parcelas = int(request.POST.get('quant_parcelas'))
                    valor_parcela = limpar_e_converter_valor(request.POST.get('valor_parcela'))
                    cartao = get_object_or_404(Cartao, id=cartao_id)

                    despesa = NotaCartao.objects.create(
                        **campos_despesa,
                        cartao=cartao,
                        quant_parcelas=quant_parcelas,
                        valor_parcela=valor_parcela,
                        status_parcelamento='a_pagar'
                    )


                elif forma_pag == 'boleto':
                    banco_id = request.POST.get('banco')
                    if banco_id:
                        banco = get_object_or_404(Banco, id=banco_id)
                    else:
                        banco = None
                    
                    vencimento = limpar_e_converter_data(request.POST.get('vencimento'), request) if request.POST.get('vencimento') else None

                    despesa = NotaBoleto.objects.create(
                        **campos_despesa,
                        recipiente=request.POST.get('recipiente'),
                        quant_boletos=request.POST.get('quant_boletos'),
                        vencimento=vencimento,
                        num_notafiscal=request.POST.get('num_notafiscal'),
                        banco=banco
                    )

                elif forma_pag == 'pix':
                    banco_id = request.POST.get('banco')
                    banco = get_object_or_404(Banco, id=banco_id)
                    despesa = NotaPix.objects.create(**campos_despesa, banco=banco)

                elif forma_pag == 'especie':
                    despesa = NotaEspecie.objects.create(
                        **campos_despesa,
                        pagador=request.POST.get('pagador')
                    )

                # **Se for obra e mão de obra, associar funcionário**
                if tipo == 'obra' and modalidade == 'mao_de_obra':
                    funcionario_id = request.POST.get('funcionario')
                    funcionario = get_object_or_404(Funcionario, id=funcionario_id)

                    MaoDeObra.objects.create(
                        despesa=despesa,
                        funcionario=funcionario,
                        categoria=request.POST.get('categoria')
                    )

                # **Redirecionamento após salvar**
                if next_url:
                    return redirect(next_url)
                return redirect('locais:detalhe_obra', id=id) if tipo == 'obra' else redirect('locais:detalhe_escritorio', escritorio_id=id)

        except IntegrityError as e:
            logger.exception(f"Erro ao criar despesa: {e}")
            messages.error(request, "Erro ao salvar a despesa. Tente novamente.")
            return redirect('locais:home', escritorio_id=escritorio_id)

    return redirect('locais:home', escritorio_id=escritorio_id)


@login_required
def editar_despesa(request, despesa_id):
    next_url = request.GET.get('next')
    despesa = get_object_or_404(Despesa, id=despesa_id)

    tipo = despesa.tipo_local.model_class().__name__.lower()  # Obtém o nome do modelo referenciado
    escritorio_id = pegar_escritorio_id(despesa, tipo)

    if request.method == 'POST':
        try:
            nome = request.POST.get('nome')
            forma_pag = request.POST.get('forma_pag')
            status = request.POST.get('status')
            valor = request.POST.get('valor')
            observacao = request.POST.get('observacao')

            if tipo == 'obra':
                modalidade = request.POST.get('modalidade')
            elif tipo == 'escritorio':
                modalidade_escritorio = request.POST.get('modalidade_escritorio')

            data = request.POST.get('data_emissao')
            data_pagamento = request.POST.get('data_pagamento')

            try:
                if data:
                    data = datetime.strptime(data, '%d/%m/%Y').date()
                    print(f"Data válida: {data} ({type(data)})")
                else:
                    print("Campo de data não preenchido")

                if data_pagamento:
                    if data_pagamento == '-':
                        data_pagamento = None
                    else:
                        data_pagamento = datetime.strptime(data_pagamento, '%d/%m/%Y').date()
                else:
                    data_pagamento = None
                    print("Campo de data_pagamento não preenchido")
            except ValueError as e:
                messages.error(request, f'Erro na data: {str(e)}')
                return redirect('locais:home', escritorio_id=escritorio_id)

            valor = limpar_e_converter_valor(valor)

            print(f"Data final antes da edição: {data} ({type(data)})")

            despesa.nome = nome
            despesa.forma_pag = forma_pag
            despesa.status = status
            despesa.valor = valor
            despesa.observacao = observacao

            if tipo == 'obra':
                despesa.modalidade = modalidade
            elif tipo == 'escritorio':
                despesa.modalidadeEscritorio = modalidade_escritorio
                
            despesa.data = data
            despesa.data_pagamento = data_pagamento
            despesa.save()

            # Tratando formas de pagamento
            if forma_pag == 'cartao':
                logger.debug("Processando pagamento com cartão...")
                cartao_id = request.POST.get('cartao')
                quant_parcelas = request.POST.get('quant_parcelas')
                valor_parcela = request.POST.get('valor_parcela')

                cartao = get_object_or_404(Cartao, id=cartao_id)
                valor_parcela = limpar_e_converter_valor(valor_parcela)

                notaCartao = get_object_or_404(NotaCartao, id=despesa_id)

                notaCartao.cartao = cartao
                notaCartao.quant_parcelas = quant_parcelas
                notaCartao.valor_parcela = valor_parcela
                notaCartao.save()
                
                logger.info(f"NotaCartao editada para despesa com cartão {cartao.id}")
            
            elif forma_pag == 'boleto':
                logger.debug("Processando pagamento com boleto...")

                vencimento = request.POST.get('vencimento')
                banco_id = request.POST.get('banco')

                if banco_id:
                    banco = get_object_or_404(Banco, id=banco_id)
                else:
                    banco = None

                vencimento = limpar_e_converter_data(vencimento, request)

                notaBoleto = get_object_or_404(NotaBoleto, id=despesa_id)

                notaBoleto.recipiente = request.POST.get('recipiente') 
                notaBoleto.quant_boletos = request.POST.get('quant_boletos') 
                notaBoleto.vencimento = vencimento
                notaBoleto.num_notafiscal = request.POST.get('num_notafiscal') 
                notaBoleto.banco = banco
                notaBoleto.save()

            elif forma_pag == 'pix':
                logger.debug("Processando pagamento com PIX...")
                banco_id = request.POST.get('banco')
                banco = get_object_or_404(Banco, id=banco_id)

                notaPix = get_object_or_404(NotaPix, id=despesa_id)
                notaPix.banco = banco
                notaPix.save()

            elif forma_pag == 'especie':
                logger.debug("Processando pagamento em espécie...")

                notaEspecie = get_object_or_404(NotaEspecie, id=despesa_id)
                notaEspecie.pagador = request.POST.get('pagador')
                notaEspecie.save()

            if tipo == 'obra':
                if modalidade == 'mao_de_obra':
                    logger.debug("Processando modalidade Mão de obra...")
                    funcionario_id = request.POST.get('funcionario')
                    funcionario = get_object_or_404(Funcionario, id=funcionario_id)

                    maoDeObra = get_object_or_404(MaoDeObra, despesa=despesa_id)
                    maoDeObra.funcionario = funcionario
                    maoDeObra.categoria = request.POST.get('categoria')
                    maoDeObra.save()

            return redirect(next_url if next_url else 'locais:home', escritorio_id=escritorio_id)

        except IntegrityError as e:
            logger.exception(f"Erro ao criar despesa: {e}")
            return redirect(next_url if next_url else 'locais:home', escritorio_id=escritorio_id)
    else:
        return redirect(next_url if next_url else 'locais:home', escritorio_id=escritorio_id)

@login_required
def deletar_despesa(request, despesa_id):
    next_url = request.GET.get('next')
    despesa = get_object_or_404(Despesa, id=despesa_id)

    despesa.delete()
    messages.success(request, 'Despesa deletada com sucesso.')
    return redirect(next_url if next_url else 'locais:home')
    
@login_required
def atualizar_status(request, despesa_id):
    next_url = request.GET.get('next')
    # Pega a despesa com o id fornecido, ou retorna 404 se não existir
    despesa = get_object_or_404(Despesa, id=despesa_id)

    if despesa.forma_pag == 'cartao':
        messages.info(request, "Operação indisponível para forma de pagamento 'Cartão'.\nPor favor, pague a fatura do cartão para alterar o status.")
    else:
        # Alterna o status da despesa entre 'a_pagar' e 'pago'
        if despesa.status == 'a_pagar':
            despesa.status = 'pago'
        else:
            despesa.status = 'a_pagar'
        
        # Salva a despesa com o novo status
        despesa.save()
    
    return redirect(next_url if next_url else 'financeiro:cartoes')


# Cartões
@login_required
def criar_cartao(request, escritorio_id):
    next_url = request.GET.get('next')

    if request.method == 'POST':
        nome = request.POST.get('nome')
        cor = request.POST.get('corCriar')
        print('COR: ', cor)
        banco_id = request.POST.get('banco')
        final = request.POST.get('final')
        vencimento = request.POST.get('vencimento')
        quant_dias = request.POST.get('quant_dias')

        # Obtenha a instância do banco ou retorne um erro 404 se não existir
        if banco_id:
            banco = get_object_or_404(Banco, id=banco_id)
        else:
            banco = None
        escritorio = get_object_or_404(Escritorio, id=escritorio_id)

        # Criando o cartão
        cartao = Cartao.objects.create(
            nome=nome,
            cor=cor,
            banco=banco,
            final=final,
            vencimento=vencimento,
            quant_dias=quant_dias,
            escritorio=escritorio
        )
        cartao.save()

    return redirect(next_url if next_url else 'financeiro:cartoes')

@login_required
def editar_cartao(request, cartao_id):
    next_url = request.GET.get('next')

    cartao = get_object_or_404(Cartao, id=cartao_id)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        cor = request.POST.get('cor')
        banco_id = request.POST.get('banco')
        final = request.POST.get('final')
        vencimento = request.POST.get('vencimento')
        quant_dias = request.POST.get('quant_dias')

        # Obtenha a instância do banco ou retorne um erro 404 se não existir
        if banco_id:
            banco = get_object_or_404(Banco, id=banco_id)
        else:
            banco = None

        cartao.nome = nome
        cartao.cor = cor
        cartao.banco = banco
        cartao.final = final
        cartao.vencimento = vencimento
        cartao.quant_dias = quant_dias

        cartao.save()
        messages.success(request, 'Cartão atualizado com sucesso.')

    return redirect(next_url if next_url else 'financeiro:cartoes')

@login_required
def deletar_cartao(request, cartao_id):
    next_url = request.GET.get('next')
    cartao = get_object_or_404(Cartao, id=cartao_id)

    cartao.delete()
    messages.success(request, 'Cartão deletado com sucesso.')
    return redirect(next_url if next_url else 'financeiro:cartoes')


@login_required
def fatura_mensal_cartoes(request):
    escritorios = Escritorio.objects.filter(membros=request.user)
    if escritorios.exists():
        escritorio_id=escritorios.first().id

    hoje = date.today()

    # Definição padrão: mês e ano atuais
    mes = hoje.month
    ano = hoje.year
    
    meses_abreviados = ["", "JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

    ano_mes = f"{ano}/{meses_abreviados[mes]}"

    if request.method == 'GET':
        ano_mes_filtro = request.GET.get('ano_mes')
        cartao_filtro = request.GET.get('cartao')

        if ano_mes_filtro:
            # Divide a string ano_mes_filtro
            ano_mes_filtro_dividido = ano_mes_filtro.split('/')

            # Extrair o ano
            ano = int(ano_mes_filtro_dividido[0])

            # Encontrar o mês na lista
            mes = 0
            for i, n in enumerate(meses_abreviados):
                if n == ano_mes_filtro_dividido[1]:
                    mes = i
                    break

            ano_mes = ano_mes_filtro

    cartoes = Cartao.objects.all()
    num_cartoes = cartoes.count()

    if cartao_filtro:
        notas_cartao = NotaCartao.objects.filter(
            status='a_pagar',
            parcelas__status='a_pagar',
            parcelas__data_vencimento__month=mes,
            parcelas__data_vencimento__year=ano,
            cartao=cartao_filtro,
        ).distinct().prefetch_related(
            Prefetch(
                'parcelas',
                queryset=Parcela.objects.filter(
                    status='a_pagar',
                    data_vencimento__month=mes,
                    data_vencimento__year=ano
                ),
                to_attr='parcela_do_mes'
            )
        ).order_by('parcelas__data_vencimento')
        cartao_selecionado = Cartao.objects.filter(id=cartao_filtro).first()

    else:
        notas_cartao = NotaCartao.objects.filter(
            status='a_pagar',
            parcelas__status='a_pagar',
            parcelas__data_vencimento__month=mes,
            parcelas__data_vencimento__year=ano,
        ).distinct().prefetch_related(
            Prefetch(
                'parcelas',
                queryset=Parcela.objects.filter(
                    status='a_pagar',
                    data_vencimento__month=mes,
                    data_vencimento__year=ano
                ),
                to_attr='parcela_do_mes'
            )
        ).order_by('parcelas__data_vencimento')
        cartao_selecionado = None

    despesas_cartao_mes = list(notas_cartao)   # Convertendo em lista para manipular diretamente

    # Calcula o total da fatura mensal
    total_fatura_mensal = 0

    for despesa in despesas_cartao_mes:
        despesa.valor_formatado = formatar_valor(despesa.valor)
        despesa.data_formatada = (despesa.data).strftime('%d/%m/%Y')

        despesa.nota_cartao = NotaCartao.objects.get(despesa_ptr_id=despesa.id)

        if despesa.parcela_do_mes:  # Verifique se existe alguma parcela para este mês
            parcela = despesa.parcela_do_mes[0]
            parcela.valor_formatado = formatar_valor(parcela.valor)
            total_fatura_mensal += parcela.valor

    
    total_fatura_mensal_formatado = formatar_valor(total_fatura_mensal)

    meses = calcular_range_meses()

    if cartoes.exists():
        for cartao in cartoes:
            if cartao.banco:
                if cartao.banco.public_id:  # Usa o public_id armazenado
                    auto_crop_url, _ = cloudinary_url(
                        cartao.banco.public_id, gravity="auto"
                    )
                    cartao.banco.imagem = auto_crop_url
                else:
                    cartao.banco.imagem = None
        
    else:
        messages.warning(request, 'Nenhum cartão cadastrado.')
    
    context = {
        'cartoes': cartoes,
        'meses': meses,
        'num_cartoes': num_cartoes,
        'despesas_cartao_mes': despesas_cartao_mes,
        'total_fatura_mensal': total_fatura_mensal_formatado,
        'ano_mes_selecionado': ano_mes if ano_mes else None,
        'cartao_selecionado': cartao_selecionado,
        'escritorio_id': escritorio_id
    }

    return render(request, 'financeiro/cartoes.html', context)


@login_required
def pagar_cartao(request, cartao_id):
    next_url = request.GET.get('next')

    # Obtém o cartão com o ID fornecido
    cartao = get_object_or_404(Cartao, id=cartao_id)

    hoje = date.today()
    mes, ano = hoje.month, hoje.year

    meses_abreviados = ["", "JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
    
    # Obtém as notas de cartão com status "a pagar"
    notas_cartao = NotaCartao.objects.filter(
        status='a_pagar',
        parcelas__status='a_pagar',
        parcelas__data_vencimento__month=mes,
        parcelas__data_vencimento__year=ano
    ).distinct().prefetch_related(
        Prefetch(
            'parcelas',
            queryset=Parcela.objects.filter(
                status='a_pagar',
                data_vencimento__month=mes,
                data_vencimento__year=ano
            ),
            to_attr='parcela_do_mes'  # Nome do atributo para acessar no template
        )
    )

    if not notas_cartao.exists():
        messages.info(request, f"Nenhuma fatura pendente para este cartão no mês atual ({meses_abreviados[mes]}/{ano}).")
        return redirect(next_url if next_url else 'financeiro:cartoes')

    total = 0

    for nota in notas_cartao:
        despesa = Despesa.objects.get(id=nota.despesa_ptr_id)
        nota.atualizar_proximo_pagamento()

        if nota.status_parcelamento == 'pago':
            despesa.status = 'pago'
            despesa.save()

        for parcela in nota.parcela_do_mes:
            if parcela.status == 'a_pagar':  # Verifica se a parcela ainda está pendente
                # Cria o pagamento
                Pagamento.objects.create(
                    parcela=parcela,
                    valor_pago=parcela.valor,
                    data_pagamento=date.today()
                )

                # Marca a parcela como paga
                parcela.status = 'pago'
                parcela.save()
                total += parcela.valor

    if request.method == 'POST':
        observacao = request.POST.get('observacao_fatura')

    Fatura.objects.create(
        cartao=cartao,
        data_pagamento=hoje,
        valor=total,
        observacao=observacao
    )
    
    messages.success(request, "Pagamentos atualizados com sucesso.")
    return redirect(next_url if next_url else 'financeiro:cartoes')



@login_required
# Editar data pagamento da parcela (que foi colocada automaticamente no pagamento da fatura)
def editar_fatura(request, fatura_id):
    next_url = request.GET.get('next')

    fatura = get_object_or_404(Fatura, id=fatura_id)
    cartao_da_fatura = fatura.cartao
    data_fatura = fatura.data_pagamento
    

    if request.method == 'POST':
        data_form = request.POST.get('data_pagamento')
        observacao = request.POST.get('observacao_fatura')
        data = limpar_e_converter_data(data_form, request)

        if data:
            notas_cartao = NotaCartao.objects.filter(
                status='pago',
                parcelas__status='pago',
                parcelas__data_vencimento__month=data_fatura.month,
                parcelas__data_vencimento__year=data_fatura.year,
                cartao=cartao_da_fatura,
            ).distinct().prefetch_related(
                Prefetch(
                    'parcelas',
                    queryset=Parcela.objects.filter(
                        status='pago',
                        data_vencimento__month=data_fatura.month,
                        data_vencimento__year=data_fatura.year
                    ),
                    to_attr='parcela_paga_do_mes'
                )
            ).order_by('parcelas__data_vencimento')

            for nota in notas_cartao:
                for parcela in nota.parcela_paga_do_mes:
                    pagamento = Pagamento.objects.get(parcela_id=parcela.id)
                    pagamento.data_pagamento=data
                    pagamento.save()

            fatura.data_pagamento = data
       
        if observacao:
            fatura.observacao = observacao

        fatura.save()

        messages.success(request, 'Fatura atualizada com sucesso.')

    return redirect(next_url if next_url else 'financeiro:cartoes')



# Bancos
@login_required
def criar_banco(request):
    next_url = request.GET.get('next')

    if request.method == 'POST':
        imagem = request.FILES.get('imagem_banco')
        nome = request.POST.get('nome')

        if imagem: 
            upload_result = cloudinary.uploader.upload(imagem)
            imagem_url = upload_result.get('secure_url')
            public_id = upload_result.get('public_id')  # Salva o ID gerado pelo Cloudinary
        else:
            imagem_url = None
            public_id = None 
    
        Banco.objects.create(nome=nome, public_id=public_id, imagem_url=imagem_url)

    return redirect(next_url if next_url else 'financeiro:cartoes')

@login_required
def editar_banco(request, banco_id):
    next_url = request.GET.get('next')
    banco = get_object_or_404(Banco, id=banco_id)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        imagem = request.FILES.get('imagem_banco_editar')

    banco.nome = nome

    if imagem: 
        upload_result = cloudinary.uploader.upload(imagem)
        imagem_url = upload_result.get('secure_url')
        public_id = upload_result.get('public_id')  # Salva o ID gerado pelo Cloudinary
        banco.imagem_url = imagem_url
        banco.public_id = public_id
    
    banco.save()

    return redirect(next_url if next_url else 'financeiro:cartoes')



@login_required
def deletar_banco(request, banco_id):
    next_url = request.GET.get('next')
    banco = get_object_or_404(Banco, id=banco_id)

    banco.delete()
    messages.success(request, 'Banco deletado com sucesso.')
    return redirect(next_url if next_url else 'locais:home')

# Funcionários
@login_required
def criar_funcionario(request):
    next_url = request.GET.get('next')

    if request.method == 'POST':
        nome = request.POST.get('nome')
        cargo = request.POST.get('cargo')

        funcionario = Funcionario.objects.create(
            nome=nome,
            cargo=cargo
        )
        funcionario.save()

    return redirect(next_url if next_url else 'financeiro:cartoes')

# TO-DO:
@login_required
def editar_funcionario(request, funcionario_id):
    pass
    
# Aditivos    
@login_required
def criar_aditivo(request, id):
    next_url = request.GET.get('next')

    if request.method == 'POST':
        nome = request.POST.get('nome')
        valor = request.POST.get('valor')
        dias = request.POST.get('dias')
        data = request.POST.get('data')
        banco_id = request.POST.get('banco')
        modalidade = request.POST.get('modalidade')
        observacao = request.POST.get('observacao')

        if modalidade == 'valor':
            dias = None
        elif modalidade == 'prazo':
            valor = None

        try:
            if data:  
                data = datetime.strptime(data, '%d/%m/%Y').date()
            if valor:   
                valor = limpar_e_converter_valor(valor)
            
        except ValueError:
            messages.error(request, 'Formato de data inválido. Use o formato dd/mm/yyyy.')
            return redirect('locais:home')

        # Obtenha a instância do banco ou retorne um erro 404 se não existir
        if banco_id:
            banco = get_object_or_404(Banco, id=banco_id)
        else:
            banco = None

        # Obtenha a instância da obra ou retorne um erro 404 se não existir
        obra = get_object_or_404(Obra, id=id)

        # Criando o aditivo
        aditivo = Aditivo.objects.create(
            nome=nome,
            valor=valor,
            dias=dias,
            data=data,
            banco=banco,
            observacao=observacao,
            modalidade=modalidade,
            obra=obra
        )
        aditivo.save()

    if next_url:
        return redirect(next_url)
    
    else:
        return render(request, 'locais/detalhe_obra.html', {
            'obra': obra,
        })

@login_required
def editar_aditivo(request, aditivo_id):
    next_url = request.GET.get('next')

    aditivo = get_object_or_404(Aditivo, id=aditivo_id)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        valor = request.POST.get('valor')
        dias = request.POST.get('dias')
        data = request.POST.get('data')
        banco_id = request.POST.get('banco')
        modalidade = request.POST.get('modalidade')
        observacao = request.POST.get('observacao')

        if modalidade == 'valor':
            dias = None
        elif modalidade == 'prazo':
            valor = None

        try:
            if data:
                data = datetime.strptime(data, '%d/%m/%Y').date()
            if valor:
                valor = limpar_e_converter_valor(valor)

        except ValueError:
            messages.error(request, 'Formato de data inválido. Use o formato dd/mm/yyyy.')
            return redirect('locais:home')

        # Obtenha a instância do banco ou retorne um erro 404 se não existir
        banco = get_object_or_404(Banco, id=banco_id)

        aditivo.nome = nome
        aditivo.valor = valor
        aditivo.dias = dias
        aditivo.data = data
        aditivo.banco = banco
        aditivo.observacao = observacao
        aditivo.modalidade = modalidade

        aditivo.save()
        messages.success(request, 'Aditivo atualizado com sucesso.')

    return redirect(next_url if next_url else 'locais:home')

@login_required
def deletar_aditivo(request, aditivo_id):
    next_url = request.GET.get('next')
    aditivo = get_object_or_404(Aditivo, id=aditivo_id)

    aditivo.delete()
    messages.success(request, 'Aditivo deletado com sucesso.')
    return redirect(next_url if next_url else 'locais:home')


# Adiantamento    
@login_required
def criar_adiantamento(request, id):
    next_url = request.GET.get('next')

    if request.method == 'POST':
        nome = request.POST.get('nome')
        valor = request.POST.get('valor')
        data = request.POST.get('data')
        banco_id = request.POST.get('banco')
        observacao = request.POST.get('observacao')

        try:
            if data:  
                data = datetime.strptime(data, '%d/%m/%Y').date()
            if valor:   
                valor = limpar_e_converter_valor(valor)
            
        except ValueError:
            messages.error(request, 'Formato de data inválido. Use o formato dd/mm/yyyy.')
            return redirect('locais:home')

        # Obtenha a instância do banco ou retorne um erro 404 se não existir
        banco = get_object_or_404(Banco, id=banco_id)

        # Obtenha a instância da obra ou retorne um erro 404 se não existir
        obra = get_object_or_404(Obra, id=id)

        # Criando o aditivo
        adiantamento = Adiantamento.objects.create(
            nome=nome,
            valor=valor,
            data=data,
            banco=banco,
            observacao=observacao,
            obra=obra
        )
        adiantamento.save()

    if next_url:
        return redirect(next_url)
    
    else:
        return render(request, 'locais/detalhe_obra.html', {
            'obra': obra,
        })

@login_required
def editar_adiantamento(request, adiantamento_id):
    next_url = request.GET.get('next')

    adiantamento = get_object_or_404(Adiantamento, id=adiantamento_id)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        valor = request.POST.get('valor')
        data = request.POST.get('data')
        banco_id = request.POST.get('banco')
        observacao = request.POST.get('observacao')

        try:
            if data:
                data = datetime.strptime(data, '%d/%m/%Y').date()
            if valor:
                valor = limpar_e_converter_valor(valor)
        except ValueError:
            messages.error(request, 'Formato de valor inválido. Use o formato correto.')
            return redirect(next_url if next_url else 'locais:home')

        # Obtenha a instância do banco ou retorne um erro 404 se não existir
        banco = get_object_or_404(Banco, id=banco_id)

        adiantamento.nome = nome
        adiantamento.valor = valor
        adiantamento.data = data
        adiantamento.banco = banco
        adiantamento.observacao = observacao

        adiantamento.save()
        messages.success(request, 'Adiantamento atualizado com sucesso.')

    return redirect(next_url if next_url else 'locais:home')

@login_required
def deletar_adiantamento(request, adiantamento_id):
    next_url = request.GET.get('next')
    adiantamento = get_object_or_404(Adiantamento, id=adiantamento_id)

    adiantamento.delete()
    messages.success(request, 'Adiantamento deletado com sucesso.')
    return redirect(next_url if next_url else 'locais:home')


# BM    
@login_required
def criar_bm(request, id):
    next_url = request.GET.get('next')

    if request.method == 'POST':
        nome = request.POST.get('nome')
        valor = request.POST.get('valor')
        data = request.POST.get('data')
        banco_id = request.POST.get('banco')
        codigo = request.POST.get('codigo')
        observacao = request.POST.get('observacao')

        try:
            if data:  
                data = datetime.strptime(data, '%d/%m/%Y').date()
            if valor:   
                valor = limpar_e_converter_valor(valor)
            
        except ValueError:
            messages.error(request, 'Formato de data inválido. Use o formato dd/mm/yyyy.')
            return redirect('locais:home')

        # Obtenha a instância do banco ou retorne um erro 404 se não existir
        banco = get_object_or_404(Banco, id=banco_id)

        # Obtenha a instância da obra ou retorne um erro 404 se não existir
        obra = get_object_or_404(Obra, id=id)

        # Criando o aditivo
        bm = BM.objects.create(
            nome=nome,
            valor=valor,
            data=data,
            banco=banco,
            codigo=codigo,
            observacao=observacao,
            obra=obra
        )
        bm.save()

    if next_url:
        return redirect(next_url)
    
    else:
        return render(request, 'locais/detalhe_obra.html', {
            'obra': obra,
        })

@login_required
def editar_bm(request, bm_id):
    next_url = request.GET.get('next')

    bm = get_object_or_404(BM, id=bm_id)

    if request.method == 'POST':
        nome = request.POST.get('nome')
        valor = request.POST.get('valor')
        data = request.POST.get('data')
        codigo = request.POST.get('codigo')
        banco_id = request.POST.get('banco')
        observacao = request.POST.get('observacao')

        try:
            if data:
                data = datetime.strptime(data, '%d/%m/%Y').date()
            if valor:
                valor = limpar_e_converter_valor(valor)
        except ValueError:
            messages.error(request, 'Formato de valor inválido. Use o formato correto.')
            return redirect(next_url if next_url else 'locais:home')

        # Obtenha a instância do banco ou retorne um erro 404 se não existir
        banco = get_object_or_404(Banco, id=banco_id)

        bm.nome = nome
        bm.valor = valor
        bm.data = data
        bm.codigo = codigo
        bm.banco = banco
        bm.observacao = observacao

        bm.save()
        messages.success(request, 'BM atualizada com sucesso.')

    return redirect(next_url if next_url else 'locais:home')

@login_required
def deletar_bm(request, bm_id):
    next_url = request.GET.get('next')
    bm = get_object_or_404(BM, id=bm_id)

    bm.delete()
    messages.success(request, 'BM deletada com sucesso.')
    return redirect(next_url if next_url else 'locais:home')


@login_required
def get_debito_mensal(request, id, ano, mes):
    meses_abreviados = ["", "JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]

    mes_formatado = 0

    for i, n in enumerate(meses_abreviados):
        if n == mes:
            mes_formatado = i
            break

    obra = Obra.objects.filter(id=id)

    debito_mensal = obra.calcular_debito_mensal(mes_formatado, ano) if obra else 0

    print(debito_mensal)

    return JsonResponse({"debito_mensal": debito_mensal})


import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def gerar_planilha(request, id, tipo):
    # Recupera a lista de despesas da sessão
    despesas_ids = request.session.get('despesas_ids', [])

    # Garante que despesas_ids é uma lista
    if not isinstance(despesas_ids, list):
        despesas_ids = []  # Garante que sempre será uma lista

    if tipo == 'obra':
        obra = get_object_or_404(Obra, id=id)
        # Filtra as despesas corretamente
        despesas = Despesa.objects.filter(id__in=despesas_ids) if despesas_ids else Despesa.objects.filter(tipo_local=ContentType.objects.get_for_model(obra), id_local=id)

    elif tipo == 'escritorio':
        escritorio = get_object_or_404(Escritorio, id=id)
        # Filtra as despesas corretamente
        despesas = Despesa.objects.filter(id__in=despesas_ids) if despesas_ids else Escritorio.objects.filter(tipo_local=ContentType.objects.get_for_model(escritorio), id_local=id)


    # Criar um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    if tipo == 'obra':
        ws.title = f"Despesas {obra.nome}"
    elif tipo == 'escritorio':
        ws.title = f"Despesas {escritorio.nome}"

    # Adicionar cabeçalhos
    headers = ["Data", "Descrição", "Valor", "Forma", "Status", "Parcelamento", "Cartão", "Nº Nota Fiscal", "Banco/Remetente",  "Observação"]

    # Definir a linha de cabeçalho
    header_row = 1
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)

        # Estilizar as células do cabeçalho
        cell.font = Font(bold=True, color="FFFFFF")  # Negrito e texto branco
        cell.fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # Cor de fundo azul
        cell.alignment = Alignment(horizontal="center", vertical="center")  # Alinhamento centralizado
        cell.border = Border(
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000"),
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000")
        )  # Bordas finas em todas as direções

    # Definir altura mínima para todas as linhas
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 30
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Adicionar as despesas à planilha
    for despesa in despesas:
        nome_cartao = '-'
        num_nota_fiscal = '-'

        despesa_status = "À pagar" if despesa.status == 'a_pagar' else "Pago"
        
        if despesa.forma_pag == 'cartao':
            nota_cartao = NotaCartao.objects.get(despesa_ptr_id=despesa.id)

            parcelamento_quant = f"{nota_cartao.quant_parcelas}x"
            
            nome_cartao = nota_cartao.cartao.nome

            remetente = nota_cartao.cartao.banco.nome

            forma_pag = "Cartão"

        elif despesa.forma_pag == 'boleto':
            nota_boleto = NotaBoleto.objects.get(despesa_ptr_id=despesa.id)
            parcelamento_quant = f"{nota_boleto.quant_boletos}x"

            num_nota_fiscal = nota_boleto.num_notafiscal

            remetente = nota_boleto.banco.nome

            forma_pag = "Boleto"
        
        else:
            parcelamento_quant = f"1x"
            
            if despesa.forma_pag == 'especie':
                nota_especie = NotaEspecie.objects.get(despesa_ptr_id=despesa.id)
                remetente = nota_especie.pagador
                forma_pag = "Espécie"
            
            elif despesa.forma_pag == 'pix':
                nota_pix = NotaPix.objects.get(despesa_ptr_id=despesa.id)
                remetente = nota_pix.banco.nome
                forma_pag = "Pix"
            

        despesa.observacao = despesa.observacao or '-'

        ws.append([
            despesa.data.strftime("%d/%m/%Y"),
            despesa.nome,
            despesa.valor,
            forma_pag,
            despesa_status,
            parcelamento_quant,
            nome_cartao,
            num_nota_fiscal,
            remetente,
            despesa.observacao,
        ])

        # Aplica formatação de moeda na coluna de valor (coluna C = índice 3)
        valor_cell = ws.cell(row=ws.max_row, column=3)
        valor_cell.number_format = 'R$ #,##0.00'  # Formato brasileiro

        # Aplicar alinhamento e quebra de linha na nova linha
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col in ws.columns:
        col_letter = col[0].column_letter  # Obtém a letra da coluna
        ws.column_dimensions[col_letter].width = 18
    
    # Criar a resposta HTTP com o arquivo
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if tipo == 'obra':
        nome_obra = obra.nome.replace(" ", "")
        response["Content-Disposition"] = f'attachment; filename="despesas_obra_{nome_obra}.xlsx"'
    elif tipo == 'escritorio':
        nome_escritorio = escritorio.nome.replace(" ", "")
        response["Content-Disposition"] = f'attachment; filename="despesas_escritorio_{nome_escritorio}.xlsx"'


    del request.session['despesas_ids']

    # Salvar a planilha no response
    wb.save(response)

    return response
